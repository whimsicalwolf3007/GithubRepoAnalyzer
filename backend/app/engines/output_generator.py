"""
Output Generator — Generates Excel reports and JSON exports
with analysis results, feasibility scores, and recommendations.
"""
import logging
from pathlib import Path
from typing import List, Dict, Any
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.config import settings

logger = logging.getLogger(__name__)

# ── Styling constants ──────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="ffffff", size=11)
BUILDABLE_FILL = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
FIXES_FILL = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
NOT_BUILDABLE_FILL = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _apply_header_style(ws, row: int, cols: int):
    """Apply header styling to a row."""
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _get_status_fill(feasibility_class: str):
    """Get fill color based on feasibility class."""
    if feasibility_class == "buildable":
        return BUILDABLE_FILL
    elif feasibility_class == "buildable_with_fixes":
        return FIXES_FILL
    else:
        return NOT_BUILDABLE_FILL


def generate_excel_report(repositories: List[Dict[str, Any]], output_path: str = None) -> str:
    """
    Generate a comprehensive Excel report with analysis results.
    
    Args:
        repositories: List of repository data dicts with analysis results
        output_path: Optional custom output path; defaults to reports directory
        
    Returns:
        Path to the generated Excel file
    """
    if not output_path:
        reports_dir = Path(settings.REPORTS_DIR)
        reports_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = str(reports_dir / f"autodev_report_{timestamp}.xlsx")

    wb = openpyxl.Workbook()

    # ── Summary Sheet ──────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # Title
    ws_summary.merge_cells("A1:H1")
    title_cell = ws_summary["A1"]
    title_cell.value = "AutoDev Intelligence — Repository Analysis Report"
    title_cell.font = Font(name="Calibri", bold=True, size=16, color="1a1a2e")
    title_cell.alignment = Alignment(horizontal="center")

    ws_summary.merge_cells("A2:H2")
    ws_summary["A2"].value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_summary["A2"].font = Font(name="Calibri", size=10, color="666666")
    ws_summary["A2"].alignment = Alignment(horizontal="center")

    # Statistics
    total = len(repositories)
    buildable = sum(1 for r in repositories if r.get("feasibility_class") == "buildable")
    fixes = sum(1 for r in repositories if r.get("feasibility_class") == "buildable_with_fixes")
    not_buildable = sum(1 for r in repositories if r.get("feasibility_class") == "not_buildable")
    avg_score = sum(r.get("feasibility_score", 0) for r in repositories) / max(total, 1)

    stats_start = 4
    stats = [
        ("Total Repositories", total),
        ("Buildable", buildable),
        ("Buildable with Fixes", fixes),
        ("Not Buildable", not_buildable),
        ("Average Feasibility Score", f"{avg_score:.1f}/100"),
    ]
    for i, (label, value) in enumerate(stats):
        ws_summary.cell(row=stats_start + i, column=1, value=label).font = Font(bold=True)
        ws_summary.cell(row=stats_start + i, column=2, value=value)

    # Repository table
    table_start = stats_start + len(stats) + 2
    headers = [
        "Repository", "Owner", "URL", "Languages", "Frameworks",
        "Feasibility Score", "Classification", "Recommendations"
    ]
    for col, header in enumerate(headers, 1):
        ws_summary.cell(row=table_start, column=col, value=header)
    _apply_header_style(ws_summary, table_start, len(headers))

    for i, repo in enumerate(repositories):
        row = table_start + 1 + i
        langs = ", ".join(repo.get("languages", {}).keys()) if isinstance(repo.get("languages"), dict) else ""
        frameworks = ", ".join(repo.get("frameworks", []))
        fc = repo.get("feasibility_class", "pending")
        rec_count = len(repo.get("recommendations", []))

        ws_summary.cell(row=row, column=1, value=repo.get("name", ""))
        ws_summary.cell(row=row, column=2, value=repo.get("owner", ""))
        ws_summary.cell(row=row, column=3, value=repo.get("url", ""))
        ws_summary.cell(row=row, column=4, value=langs)
        ws_summary.cell(row=row, column=5, value=frameworks)
        ws_summary.cell(row=row, column=6, value=repo.get("feasibility_score", 0))
        
        class_cell = ws_summary.cell(row=row, column=7, value=fc.replace("_", " ").title())
        class_cell.fill = _get_status_fill(fc)
        
        ws_summary.cell(row=row, column=8, value=f"{rec_count} recommendations")

        # Apply borders
        for col in range(1, len(headers) + 1):
            ws_summary.cell(row=row, column=col).border = THIN_BORDER

    # Auto-fit column widths
    for col in range(1, len(headers) + 1):
        ws_summary.column_dimensions[get_column_letter(col)].width = 20

    # ── Recommendations Sheet ──────────────────────────────────
    ws_recs = wb.create_sheet("Recommendations")
    rec_headers = [
        "Repository", "Category", "Severity", "Title",
        "Description", "Fix", "Effort", "Estimated Time", "AI Provider"
    ]
    for col, header in enumerate(rec_headers, 1):
        ws_recs.cell(row=1, column=col, value=header)
    _apply_header_style(ws_recs, 1, len(rec_headers))

    rec_row = 2
    for repo in repositories:
        for rec in repo.get("recommendations", []):
            ws_recs.cell(row=rec_row, column=1, value=repo.get("name", ""))
            ws_recs.cell(row=rec_row, column=2, value=rec.get("category", ""))
            ws_recs.cell(row=rec_row, column=3, value=rec.get("severity", ""))
            ws_recs.cell(row=rec_row, column=4, value=rec.get("title", ""))
            ws_recs.cell(row=rec_row, column=5, value=rec.get("description", ""))
            ws_recs.cell(row=rec_row, column=6, value=rec.get("fix", ""))
            ws_recs.cell(row=rec_row, column=7, value=rec.get("effort", ""))
            ws_recs.cell(row=rec_row, column=8, value=rec.get("estimated_time", ""))
            ws_recs.cell(row=rec_row, column=9, value=rec.get("ai_provider", ""))
            for col in range(1, len(rec_headers) + 1):
                ws_recs.cell(row=rec_row, column=col).border = THIN_BORDER
            rec_row += 1

    for col in range(1, len(rec_headers) + 1):
        ws_recs.column_dimensions[get_column_letter(col)].width = 22

    # Save
    wb.save(output_path)
    logger.info(f"Excel report saved to: {output_path}")
    return output_path
